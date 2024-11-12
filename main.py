import cv2
import os
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from tqdm import tqdm  # 导入 tqdm 库
from datetime import datetime  # 导入 datetime 模块
import subprocess  # 导入 subprocess，用于调用 ffmpeg

# 定义分辨率范围（宽度和高度）
RESOLUTION_RANGES = {
    "720P": ((1280, 1366), (680, 800)),
    "1080P": ((1920, 2048), (1000, 1200)),
    "4K": ((3840, 4096), (2160, 2304))
}

# 定义文件路径
input_folder = r"\\Win-rdfn9isbbrj\e\美日韩电影"

# 获取当前日期
current_date = datetime.now().strftime("%Y-%m-%d")  # 获取当前日期并格式化为字符串

# 自动生成包含日期的输出文件路径
output_file = f"C:\\更新信息_{current_date}.xlsx"  # 输出 Excel 文件路径，添加日期

# 用来存储结果
results = []

# 获取所有子文件夹
folders = [folder_name for folder_name in os.listdir(input_folder) if os.path.isdir(os.path.join(input_folder, folder_name))]

# 使用 tqdm 昿示进度条
for folder_name in tqdm(folders, desc="Processing folders", ncols=100):  # 添加进度条
    folder_path = os.path.join(input_folder, folder_name)
    found_resolutions = set()  # 存储当前文件夹中的分辨率
    mkv_or_mp4_count = 0  # 用于计数 .mkv 和 .mp4 文件数量

    # 遍历子文件夹及其文件
    for root, dirs, files in os.walk(folder_path):
        for file_name in files:
            file_path = os.path.join(root, file_name)

            # 检查文件扩展名，确保是视频文件
            if file_name.lower().endswith(('.mp4', '.mkv', '.avi', '.mov', '.flv', '.webm', '.mpg', '.mpeg')):
                # 计数 .mkv 或 .mp4 文件数量
                if file_name.lower().endswith(('.mkv', '.mp4')):
                    mkv_or_mp4_count += 1

                # 检查文件名中是否包含分辨率信息（例如 -1080P 或 -4K）
                if "-1080P" in file_name.upper():
                    found_resolutions.add("1080P")
                elif "-4K" in file_name.upper():
                    found_resolutions.add("4K")
                elif "-720P" in file_name.upper():
                    found_resolutions.add("720P")

                # 如果文件名没有分辨率信息，尝试使用 cv2 获取分辨率
                if not found_resolutions:
                    try:
                        cap = cv2.VideoCapture(file_path)
                        if cap.isOpened():
                            width = cap.get(cv2.CAP_PROP_FRAME_WIDTH)
                            height = cap.get(cv2.CAP_PROP_FRAME_HEIGHT)
                            cap.release()

                            # 根据宽度和高度范围判断分辨率
                            for resolution, ((min_width, max_width), (min_height, max_height)) in RESOLUTION_RANGES.items():
                                if min_width <= width <= max_width and min_height <= height <= max_height:
                                    found_resolutions.add(resolution)
                                    break  # 找到满足条件的分辨率即可退出
                    except Exception as e:
                        # 只在无法处理视频时打印错误，避免输出太多信息
                        pass

                    # 如果 cv2 获取不到分辨率，使用 ffmpeg 获取分辨率
                    if not found_resolutions:
                        try:
                            result = subprocess.run(
                                ["ffmpeg", "-i", file_path],
                                stderr=subprocess.PIPE,  # 获取错误输出
                                stdout=subprocess.PIPE,
                                universal_newlines=True
                            )
                            # 通过解析 stderr 中的输出获取分辨率
                            for line in result.stderr.splitlines():
                                if "Video:" in line and "x" in line:  # 查找分辨率信息
                                    resolution_str = line.split(",")[2].strip()
                                    width, height = map(int, resolution_str.split("x"))
                                    # 根据宽度和高度范围判断分辨率
                                    for resolution, ((min_width, max_width), (min_height, max_height)) in RESOLUTION_RANGES.items():
                                        if min_width <= width <= max_width and min_height <= height <= max_height:
                                            found_resolutions.add(resolution)
                                            break
                                    break  # 找到后退出
                        except Exception as e:
                            # 只在无法处理视频时打印错误，避免输出太多信息
                            pass

    # 如果未识别到信息，则替换为 .mkv 或 .mp4 文件的总数量
    resolutions_status = [res for res in RESOLUTION_RANGES if res in found_resolutions]
    if not resolutions_status:
        resolutions_status.append(f"总文件数: {mkv_or_mp4_count}")  # 替换为文件夹下 MKV 或 MP4 文件的总数量

    # 将结果添加到列表
    results.append([folder_name, ", ".join(resolutions_status), "已存在"])

# 使用 pandas 将结果转换为 DataFrame
df = pd.DataFrame(results, columns=["影视名称", "视频分辨率", "存在状态"])

# 使用 openpyxl 创建一个工作簿
wb = Workbook()
ws = wb.active

# 将 DataFrame 的内容添加到工作簿
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# 自动调整列宽并居中对齐
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # 获取列字母
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)  # 根据内容长度来设置列宽，适当增加 2
    ws.column_dimensions[column].width = adjusted_width

# 手动设置 "影视名称" 列宽为 20
ws.column_dimensions['A'].width = 20  # 'A' 列即 "影视名称" 列
ws.column_dimensions['C'].width = 10

# 设置单元格内容居中
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

# 设置标题行加粗并居中
for cell in ws[1]:
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold=True)

# 设置表格内容的字体（可选，增加美观）
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.font = Font(name='Arial', size=10)

# 保存到 Excel 文件
wb.save(output_file)

print(f"文件信息已成功保存为 Excel 文件：{output_file}")
